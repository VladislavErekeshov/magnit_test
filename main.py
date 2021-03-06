import sqlite3 as sql
import openpyxl
import xlsxwriter
import pdfminer.high_level
import sys
import fpdf

# создаём базу данных и подключаемся к ней
conn = sql.connect('database.db')
cur = conn.cursor()

# функция для создания таблиц в бд
def create_tables():
    # создаём таблицу users
    cur.execute("""CREATE TABLE IF NOT EXISTS users(
        id INTEGER NOT NULL PRIMARY KEY,
        second_name TEXT NOT NULL,
        first_name TEXT NOT NULL,
        patronymic TEXT,
        region_id INTEGER NOT NULL,
        city_id INTEGER NOT NULL,
        phone TEXT NOT NULL,
        email TEXT);
        """)

    # создаём таблицу regions
    cur.execute("""CREATE TABLE IF NOT EXISTS regions(
        id INTEGER PRIMARY KEY NOT NULL,
        region_name TEXT NOT NULL);
        """)

    # создаём таблицу cities
    cur.execute("""CREATE TABLE IF NOT EXISTS cities(
        id INTEGER PRIMARY KEY NOT NULL,
        region_id INTEGER NOT NULL,
        city_name TEXT NOT NULL);
        """)

    # заполняем первичные данные
    regions = [(0, "Краснодарский край"), (1, "Ростовская область"), (2, "Ставропольский край")]
    cur.executemany("INSERT OR IGNORE INTO regions VALUES(?, ?);", regions)

    cities = [(0, 0, "Краснодар"), (1, 0, "Кропоткин"), (2, 0, "Славянск"), (3, 1, "Ростов"), (4, 1, "Шахты"), (5, 1, "Батайск"), (6, 2, "Ставрополь"), (7, 2, "Пятигорск"), (8, 2, "Кисловодск")]
    cur.executemany("INSERT OR IGNORE INTO cities VALUES(?, ?, ?);", cities)

    conn.commit()


# функция, которая импортирует данные из Excel в таблицу users
def xlsx_import():
    try:
        try:
            wb = openpyxl.reader.excel.load_workbook(filename = "import.xlsx", data_only=True) # читаем Excel файл, из которого нам нужно импортировать данные
        except FileNotFoundError as e:
            print(e)
            sys.exit(1)

        # выбираем активный лист
        wb.active = 0
        sheet = wb.active

        # парсим xlsx документ и заполняем таблицу users
        for i in range(2, 1048577): # по версии Google в листе Excel содержится 1048576 строк
            if sheet['A' + str(i)].value == None: # завершаем цикл когда программа наткнется на пустую ячейку А
                break
            else:
                # находим id региона с помощью таблицы regions и сохраняем в переменную
                region_name = sheet['E' + str(i)].value # парсим название региона
                if region_name != "Краснодарский край" and region_name != "Ростовская область" and region_name != "Ставропольский край":
                    print("Ошибка в названии региона юзера с id '{0}'".format(sheet['A' + str(i)].value))
                    sys.exit(1)
                else:
                    cur.execute('SELECT id FROM regions WHERE region_name = ?', (region_name,))
                    region_id = cur.fetchall()

                # находим id города с помощью таблицы cities и сохраняем в переменную
                city_name = sheet['F' + str(i)].value  # парсим название города
                if city_name != "Краснодар" and city_name != "Кропоткин" and city_name != "Славянск" and city_name != "Ростов" and city_name != "Шахты" and city_name != "Батайск" and city_name != "Ставрополь" and city_name != "Пятигорск" and city_name != "Кисловодск":
                    print("Ошибка в названии города юзера с id '{0}'".format(sheet['A' + str(i)].value))
                    sys.exit(1)
                else:
                    cur.execute('SELECT region_id FROM cities WHERE city_name = ?', (city_name,))
                    region_city_id = cur.fetchall()
                    if region_city_id != region_id:
                        print("Город юзера с id '{0}' не может находиться в этом регионе".format(sheet['A' + str(i)].value))
                        sys.exit(1)
                    else:
                        cur.execute('SELECT id FROM cities WHERE city_name = ?', (city_name,))
                        city_id = cur.fetchall()

                # сохраняем спаршенные данные одной строки в кортеж и кидаем в таблицу users
                values = (sheet['A' + str(i)].value, sheet['B' + str(i)].value, sheet['C' + str(i)].value, sheet['D' + str(i)].value, region_id[0][0], city_id[0][0], sheet['G' + str(i)].value, sheet['H' + str(i)].value)
                cur.execute("INSERT OR IGNORE INTO users VALUES(?, ?, ?, ?, ?, ?, ?, ?);", values)

        conn.commit()
    except sql.OperationalError as e:
        print(e)
        sys.exit(1) 


# функция для нахождения количества строк в таблице users
def users_len():
    try:
        cur.execute("SELECT * FROM users")
        rows = cur.fetchall()
        return len(rows)
    except sql.OperationalError as e:
        print(e)
        sys.exit(1)


# функция, которая экспортирует данные из таблицы users в Excel
def xlsx_export():
    try:
        if users_len() == 0:
            print("таблица users пустая")
            sys.exit(1)


        # создаем Excel файл
        workbook = xlsxwriter.Workbook('export.xlsx')
        worksheet = workbook.add_worksheet()

        # заполняем первую строку с наименованиями
        worksheet.write(0, 0, "id")
        worksheet.write(0, 1, "Фамилия")
        worksheet.write(0, 2, "Имя")
        worksheet.write(0, 3, "Отчество")
        worksheet.write(0, 4, "Регион")
        worksheet.write(0, 5, "Город")
        worksheet.write(0, 6, "Контактный телефон")
        worksheet.write(0, 7, "E-mail")

        # берем данные из таблицы users
        cur.execute("SELECT * FROM users")
        data = cur.fetchall()

        # заполняем Excel файл
        len = users_len()
        for i in range(0, len):
            d = data[i]  # данные i-ой строки таблицы users

            # находим название региона из таблицы regions по его id и записываем в Excel файл
            region_id =  d[4]
            cur.execute('SELECT region_name FROM regions WHERE id = ?', (region_id,))
            region_name = cur.fetchall()
            worksheet.write(i+1, 4, region_name[0][0])

            # находим название города из таблицы cities по его id и записываем в Excel файл
            city_id = d[5]
            cur.execute('SELECT city_name FROM cities WHERE id = ?', (city_id,))
            city_name = cur.fetchall()
            worksheet.write(i+1, 5, city_name[0][0])

            # записываем остальные данные
            worksheet.write(i+1, 0, d[0])
            worksheet.write(i+1, 1, d[1])
            worksheet.write(i+1, 2, d[2])
            worksheet.write(i+1, 3, d[3])
            worksheet.write(i+1, 6, d[6])
            worksheet.write(i+1, 7, d[7])

        workbook.close()
    except sql.OperationalError as e:
        print(e)
        sys.exit(1)


# функция, которая импортирует данные из pdf в таблицу users
def pdf_import():
    try:
        with open('import.pdf', 'rb') as file: # открываем pdf файл
            # парсим текст и сохраняем в переменные
            text = pdfminer.high_level.extract_text(file, sys.stdout)
            id = 0
            second_name = text[0:8]
            first_name = text[9:18]
            patronymic = text[19:32]
            phone = text[75:91]
            email = text[92:119]
            city_name = text[213:222]
            

            # находим id региона с помощью таблицы cities и сохраняем в переменную
            cur.execute('SELECT region_id FROM cities WHERE city_name = ?', (city_name,))
            region_id = cur.fetchall()

            # находим id города с помощью таблицы cities и сохраняем в переменную
            cur.execute('SELECT id FROM cities WHERE city_name = ?', (city_name,))
            city_id = cur.fetchall()

            # заполняем данными таблицу users
            values = (id, second_name, first_name, patronymic, region_id[0][0], city_id[0][0], phone, email)
            cur.execute("INSERT OR IGNORE INTO users VALUES(?, ?, ?, ?, ?, ?, ?, ?);", values)

            conn.commit()
    except (sql.OperationalError, FileNotFoundError) as e:
        print(e)
        sys.exit(1)
        

# функция, которая экспортирует данные из таблицы users в pdf файл
def pdf_export():
    try:
        # создаем pdf файл
        pdf = fpdf.FPDF(format='letter')
        pdf.add_page() 

        # выбираем и настраиваем unicode шрифт
        font_path = 'DejaVuSansCondensed.ttf'
        font_family = 'family'
        pdf.add_font(family=font_family, fname=font_path, uni=True)
        pdf.set_font(family=font_family, size=8)

        # берем данные из таблицы users
        cur.execute("SELECT * FROM users")
        data = cur.fetchall()

        # заполняем pdf файл
        len = users_len()
        for i in range(0, len):
            d = data[i] # данные i-ой строки таблицы users

            # находим название региона из таблицы regions по его id и сохраняем в переменную
            region_id =  d[4]
            cur.execute('SELECT region_name FROM regions WHERE id = ?', (region_id,))
            region_name = cur.fetchall()

            # находим название города из таблицы cities по его id и сохраняем в переменную
            city_id = d[5]
            cur.execute('SELECT city_name FROM cities WHERE id = ?', (city_id,))
            city_name = cur.fetchall()

            # прописываем варианты переменных с пустым отчествои или E-mail
            if d[3] == None:
                patronymic = ''
            else:
                patronymic = d[3]
            if d[7] == None:
                email = ''
            else:
                email = d[7]

            # заполняем pdf файл
            pdf.cell(200, 10, txt="{0} {1} {2} {3} {4} {5} {6} {7}".format(d[0], d[1], d[2], patronymic, region_name[0][0], city_name[0][0], d[6], email), ln=i+1, align="L")

        pdf.output("export.pdf")
    except sql.OperationalError as e:
        print(e)
        sys.exit(1)

# импортирующие функции друг с другом конфликтуют из-за того, что я не стал реализовать автозаполнение id в таблице (потому что не смог найти решение,
# при котором таблица не будет дублировать сама себя при повторном запуске программы), в результате получается, что корректнее всего программа будет работать
# если одна из этих функций отключена, но так же программа будет работать и при обоих включенных функциях, но в зависимости от положения функций в очереди, 
# таблица будет заполняться по разному: если на первом месте будет функция на импорт экселя, то не будет разницы в том, отключена ли функция на импорт пдф или нет, 
# результат будет один и тот же, если же первее будет функция на импорт пдф, то id 0 будет моё резюме, а остальные будут из экселя (кроме id 0)

create_tables()
xlsx_import()
xlsx_export()

#pdf_import()
pdf_export()